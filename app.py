import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os

class SistemaAsistencia:
    def __init__(self, archivo="asistencia.xlsx"):
        # Verificar si el archivo existe, si no, crearlo con la estructura inicial
        if not os.path.exists(archivo):
            self.crear_archivo_inicial(archivo)
        self.archivo = archivo
        self.workbook = openpyxl.load_workbook(archivo)
        self.sheet = self.workbook["Asistencia"]

    def crear_archivo_inicial(self, archivo):
        # Crear un nuevo libro de trabajo y hoja
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Asistencia"

        # Definir encabezados
        headers = ["Nombre", "Fecha", "Hora Entrada", "Hora Salida", "Duración (horas)"]
        sheet.append(headers)

        # Guardar el archivo
        wb.save(archivo)

    def registrar_entrada(self):
        # Solicitar información al usuario
        nombre = input("Ingrese el nombre: ").strip()
        fecha = datetime.now().strftime("%Y-%m-%d")
        hora_entrada = datetime.now().strftime("%H:%M:%S")

        # Agregar registro al Excel
        self.sheet.append([nombre, fecha, hora_entrada, "", ""])
        self.workbook.save(self.archivo)
        print(f"Entrada registrada para {nombre} a las {hora_entrada}")

    def registrar_salida(self):
        nombre = input("Ingrese el nombre: ").strip()
        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        hora_salida = datetime.now().strftime("%H:%M:%S")

        # Buscar el último registro del usuario en la fecha actual
        for row in range(2, self.sheet.max_row + 1):
            if (self.sheet.cell(row=row, column=1).value == nombre and 
                self.sheet.cell(row=row, column=2).value == fecha_actual and 
                self.sheet.cell(row=row, column=4).value == ""):

                # Registrar hora de salida
                self.sheet.cell(row=row, column=4).value = hora_salida

                # Calcular duración
                hora_entrada = datetime.strptime(self.sheet.cell(row=row, column=3).value, "%H:%M:%S")
                hora_salida_dt = datetime.strptime(hora_salida, "%H:%M:%S")
                duracion = (hora_salida_dt - hora_entrada).total_seconds() / 3600  # En horas
                self.sheet.cell(row=row, column=5).value = round(duracion, 2)

                self.workbook.save(self.archivo)
                print(f"Salida registrada para {nombre}. Duración: {duracion:.2f} horas")
                return

        print("No se encontró un registro de entrada para hoy.")

    def generar_reporte_individual(self, nombre):
        # Leer datos con pandas
        df = pd.read_excel(self.archivo)

        # Filtrar por nombre
        reporte = df[df["Nombre"] == nombre]

        if reporte.empty:
            print(f"No hay registros para {nombre}")
        else:
            print(f"\nReporte de asistencia para {nombre}:")
            print(reporte.to_string(index=False))
            print(f"Horas totales: {reporte['Duración (horas)'].sum():.2f}")

    def visualizar_estadisticas(self):
        # Leer datos con pandas
        df = pd.read_excel(self.archivo)

        # Agrupar por nombre y sumar horas
        estadisticas = df.groupby("Nombre")["Duración (horas)"].sum().dropna()

        if estadisticas.empty:
            print("No hay datos para mostrar estadísticas.")
            return

        # Crear gráfico de barras
        plt.figure(figsize=(10, 6))
        estadisticas.plot(kind="bar")
        plt.title("Horas Totales por Persona")
        plt.xlabel("Nombre")
        plt.ylabel("Horas")
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.show()

def menu_principal():
    sistema = SistemaAsistencia()
    while True:
        print("\n===== SISTEMA DE GESTIÓN DE ASISTENCIA =====")
        print("1. Registrar entrada")
        print("2. Registrar salida")
        print("3. Generar reporte individual")
        print("4. Visualizar estadísticas")
        print("5. Salir")

        opcion = input("\nSeleccione una opción: ")

        if opcion == "1":
            sistema.registrar_entrada()
        elif opcion == "2":
            sistema.registrar_salida()
        elif opcion == "3":
            nombre = input("Ingrese el nombre para el reporte: ").strip()
            sistema.generar_reporte_individual(nombre)
        elif opcion == "4":
            sistema.visualizar_estadisticas()
        elif opcion == "5":
            print("¡Hasta luego!")
            break
        else:
            print("Opción no válida. Intente de nuevo.")

if __name__ == "__main__":
    menu_principal()